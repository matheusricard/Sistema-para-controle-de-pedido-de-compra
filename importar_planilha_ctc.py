import os
import sqlite3
from datetime import datetime

from openpyxl import load_workbook  # pip install openpyxl

# Caminhos baseados na pasta onde este script está
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(BASE_DIR, "sistema_compras.db")

# Coloque aqui o NOME EXATO da sua planilha
XLSX_FILE = os.path.join(BASE_DIR, "ctc_compras.xlsx")



def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Garante que a tabela pedidos exista e que tenha as colunas novas."""
    conn = get_db_connection()
    cur = conn.cursor()

    # Cria a tabela se não existir
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_veiculo TEXT,
            tag TEXT,
            numero_sc TEXT,
            data_criacao_sc TEXT,
            numero_pc TEXT,
            descricao_itens TEXT,
            status_pedido TEXT,
            data_pagamento TEXT,
            numero_nota_fiscal TEXT,
            valor_pedido REAL,
            nome_fornecedor TEXT,
            local TEXT,
            obra TEXT,
            entrega_financeiro TEXT,
            departamento TEXT,
            observacao TEXT,
            data_cadastro TEXT
        )
        """
    )

    # Se a tabela já existia sem as colunas novas, tenta adicionar via ALTER TABLE
    try:
        cur.execute("ALTER TABLE pedidos ADD COLUMN entrega_financeiro TEXT")
    except sqlite3.OperationalError:
        # coluna já existe, ignora
        pass

    try:
        cur.execute("ALTER TABLE pedidos ADD COLUMN departamento TEXT")
    except sqlite3.OperationalError:
        # coluna já existe, ignora
        pass

    conn.commit()
    conn.close()


def parse_valor_import(valor):
    """Converte valor da planilha em float (aceita número ou string com vírgula)."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if not s:
        return 0.0
    # Tira separador de milhar e troca vírgula por ponto
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_data(cell):
    """Converte datas da planilha para 'AAAA-MM-DD' quando possível."""
    if cell is None:
        return ""
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    s = str(cell).strip()
    if not s:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Se não bater nenhum formato, devolve a string original
    return s


def ja_existe_pedido(cur, numero_pc, numero_sc):
    """
    Trava para não duplicar pedido.
    Considera duplicado se tiver mesmo PC e mesma SC.
    """
    if not numero_pc and not numero_sc:
        return False
    cur.execute(
        """
        SELECT id FROM pedidos
        WHERE numero_pc = ? AND numero_sc = ?
        """,
        (numero_pc, numero_sc),
    )
    return cur.fetchone() is not None


def importar():
    if not os.path.exists(XLSX_FILE):
        raise FileNotFoundError(f"Planilha não encontrada em: {XLSX_FILE}")

    print(f"Lendo planilha: {XLSX_FILE}")
    wb = load_workbook(XLSX_FILE, data_only=True)

    # Aba onde estão os dados (pelo que você me mandou, é COMPRAS)
    if "COMPRAS" not in wb.sheetnames:
        raise ValueError("A planilha não tem uma aba chamada 'COMPRAS'.")
    sheet = wb["COMPRAS"]

    conn = get_db_connection()
    cur = conn.cursor()

    inseridos = 0
    pulados = 0

    # Cabeçalho está na linha 3, dados a partir da 4
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        cols = list(row)

        # Ignora linhas completamente vazias nas primeiras colunas
        if not any(cols[:16]):
            continue

        # Mapeamento das colunas da planilha para os campos do sistema:
        # 1  -> EQUIP / DEPART       -> nome_veiculo
        # 2  -> TAG                  -> tag
        # 3  -> DESCRIÇÃO            -> descricao_itens
        # 4  -> DATA DA SC           -> data_criacao_sc
        # 5  -> SOLICITAÇÃO          -> numero_sc
        # 6  -> PEDIDO DE COMPRA     -> numero_pc
        # 7  -> FORNECEDOR           -> nome_fornecedor
        # 8  -> STATUS DO PEDIDO     -> status_pedido
        # 9  -> DATA DO PAGAMENTO    -> data_pagamento
        # 10 -> DATA ENTREGA / NF    -> numero_nota_fiscal (texto completo)
        # 11 -> VALOR DO PEDIDO      -> valor_pedido
        # 12 -> CIDADE               -> local
        # 13 -> ENTREGA NO FINANCEIRO-> entrega_financeiro (COLUNA NOVA NO BD)
        # 14 -> OBSERVAÇÃO RELEVANTE -> observacao
        # 15 -> OBRA                 -> obra
        # 16 -> DEPART.              -> departamento (COLUNA NOVA NO BD)

        def s(v):
            if v is None:
                return ""
            return str(v).strip()

        nome_veiculo = s(cols[0] if len(cols) > 0 else "")
        tag = s(cols[1] if len(cols) > 1 else "")
        descricao_itens = s(cols[2] if len(cols) > 2 else "")

        data_criacao_sc = parse_data(cols[3] if len(cols) > 3 else None)

        numero_sc = s(cols[4] if len(cols) > 4 else "")
        numero_pc = s(cols[5] if len(cols) > 5 else "")

        nome_fornecedor = s(cols[6] if len(cols) > 6 else "")
        status_pedido = s(cols[7] if len(cols) > 7 else "")

        data_pagamento = parse_data(cols[8] if len(cols) > 8 else None)

        entrega_nf = s(cols[9] if len(cols) > 9 else "")
        numero_nota_fiscal = entrega_nf  # texto completo (data + NF, se vier assim)

        valor_pedido = parse_valor_import(cols[10] if len(cols) > 10 else None)

        local = s(cols[11] if len(cols) > 11 else "")

        entrega_financeiro = s(cols[12] if len(cols) > 12 else "")
        observacao = s(cols[13] if len(cols) > 13 else "")
        obra = s(cols[14] if len(cols) > 14 else "")
        departamento = s(cols[15] if len(cols) > 15 else "")

        data_cadastro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Evita duplicar SC/PC
        if ja_existe_pedido(cur, numero_pc, numero_sc):
            pulados += 1
            continue

        cur.execute(
            """
            INSERT INTO pedidos (
                nome_veiculo, tag, numero_sc, data_criacao_sc, numero_pc,
                descricao_itens, status_pedido, data_pagamento, numero_nota_fiscal,
                valor_pedido, nome_fornecedor, local, obra,
                entrega_financeiro, departamento, observacao, data_cadastro
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                nome_veiculo,
                tag,
                numero_sc,
                data_criacao_sc,
                numero_pc,
                descricao_itens,
                status_pedido,
                data_pagamento,
                numero_nota_fiscal,
                valor_pedido,
                nome_fornecedor,
                local,
                obra,
                entrega_financeiro,
                departamento,
                observacao,
                data_cadastro,
            ),
        )
        inseridos += 1

    conn.commit()
    conn.close()

    print("Importação concluída.")
    print(f"Linhas inseridas: {inseridos}")
    print(f"Linhas puladas como duplicadas: {pulados}")


if __name__ == "__main__":
    print("Iniciando importação da planilha CTC para o banco sistema_compras.db...")
    init_db()
    importar()
